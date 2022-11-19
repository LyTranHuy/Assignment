import csv
import json

# Convert json to csv

"""with open('input_test.json') as json_file:
    data = json.load(json_file)
data_input = data['thecat']

data_file = open('output_test.csv','w')

csv_writer = csv.writer(data_file)


count = 0

for item in data_input:
    if count == 0:
        header = data_input.keys()
        csv_writer.writerow(header)
        csv_writer.writerow(data_input.values())
        count += 1
data_file.close()    """

# Convert csv to json

"""import csv
import json

jsonArray=[]

#read csv file
with open('input_test.csv') as csvfile:
    #load csv file using csv library's dictreader method
    csvRead = csv.DictReader(csvfile)

    #convert row csv row to python dictionary
    for row in csvRead:
        jsonArray.append(row)

#convert python jsonarray to json string and write in into a file
with open('out_test.json','w') as jsonfile:
    jsonString = json.dumps(jsonArray, indent=4)
    jsonfile.write(jsonString)
"""

# Convert json to xlxs


import json
from openpyxl import Workbook

keys = []
wb = Workbook()
ws = wb.active

with open('input_test.json') as f:
    json_file = json.load(f)
	
json_data = json_file['thecat']
for i in range(len(json_data)) :
	sub_obj = json_data[i]
	if i == 0 :
		keys = list(sub_obj.keys())
		for k in range(len(keys)) :
			# row or column index start from 1
			ws.cell(row = (i + 1), column = (k + 1), value = keys[k]);
	for j in range(len(keys)) :
		ws.cell(row = (i + 2), column = (j + 1), value = sub_obj[keys[j]]);
wb.save("output_test.xlsx")



